"""Build a single Excel workbook that surfaces every text field the
workbench renders for every L2 question, plus the value-stream tree,
the process-flow data, and the catalogue of UI default/placeholder
strings the cards display.

Use case: walk through each row and confirm the data → UI mapping is
correct end-to-end (verbatim, plain-language, consultant wording,
aboutText, standardMeans, fioriApps, etc.).

Run from repo root:

    python3 scripts/build-content-audit.py

Output: content-audit.xlsx in the repo root.

Sheets:
  1. Questions     — 150 rows, every text field per question
  2. Scope items   — 672 rows
  3. Sub-processes — 56 rows
  4. Value streams — 8 rows
  5. Process flow  — 2,502 step rows (SAP-verbatim activity + Fiori)
  6. Per-scope flow summary — 655 rows (one per scope item with flow)
  7. UI defaults   — every placeholder + default label the cards show
"""

from __future__ import annotations

import json
import os
from collections import Counter
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATASET = os.path.join(REPO, "prisma", "seeds", "value-stream", "dataset.json")
PROCESS_FLOW = os.path.join(REPO, "prisma", "seeds", "value-stream", "process-flow.json")
OUT = os.path.join(REPO, "content-audit.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="002B5C")
HEADER_FONT = Font(bold=True, color="FFFFFE", name="Inter")
EYEBROW_FILL = PatternFill("solid", fgColor="F4F2EB")
WRAP = Alignment(wrap_text=True, vertical="top")


def write_header(ws, row: int, columns: Iterable[str]) -> None:
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, widths: list[tuple[int, int]]) -> None:
    for col, width in widths:
        ws.column_dimensions[get_column_letter(col)].width = width


def short(s: object, n: int = 80) -> str:
    if s is None:
        return ""
    t = str(s)
    return t if len(t) <= n else t[: n - 1] + "…"


# ─── Load data ──────────────────────────────────────────────────────

with open(DATASET, "r", encoding="utf-8") as fh:
    ds = json.load(fh)
with open(PROCESS_FLOW, "r", encoding="utf-8") as fh:
    pf = json.load(fh)

# Index helpers
sub_by_id = {s["id"]: s for s in ds["subProcesses"]}
stream_by_id = {s["id"]: s for s in ds["valueStreams"]}
scope_by_id = {s["id"]: s for s in ds["scopeItems"]}
flow_by_id = {f["scopeItemId"]: f for f in pf["flows"]}
steps_by_scope: dict[str, list[dict]] = {}
for s in pf["steps"]:
    steps_by_scope.setdefault(s["scopeItemId"], []).append(s)
for arr in steps_by_scope.values():
    arr.sort(key=lambda r: r["stepNumber"])
questions_by_scope_ref: dict[str, list[str]] = {}
questions_by_sub: dict[str, list[str]] = {}
for q in ds["questions"]:
    questions_by_sub.setdefault(q["subProcessId"], []).append(q["id"])
    for ref in q["scopeItemRefs"]:
        questions_by_scope_ref.setdefault(ref, []).append(q["id"])

wb = Workbook()
wb.remove(wb.active)


# ─── Sheet 1: Questions (the headline audit table) ──────────────────

ws = wb.create_sheet("Questions")
columns = [
    "#",
    "Question ID",
    "Status",
    "Format default",
    "Flag",
    "Value stream",
    "Sub-process",
    "Scope-item refs",
    "Scope-item count",
    "SAP Area",
    "SAP Topic",
    "SSCUI",
    "Source BDC",
    "Placement basis",
    "SAP verbatim (LOCKED — never edited)",
    "Plain-language suggested (consultant draft)",
    "Consultant wording (editable, what the client sees)",
    "About text — SAP Topic Definition (seeds 'What this is about')",
    "Standard means (consultant fills, the teal box)",
    "Status note",
    "Is custom (consultant-added)",
]
write_header(ws, 1, columns)
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

# Sort by stream displayOrder, then sub-process displayOrder, then displayOrder.
def stream_order(sid: str) -> int:
    return stream_by_id.get(sid, {}).get("displayOrder", 99)


def sub_order(spid: str) -> int:
    return sub_by_id.get(spid, {}).get("displayOrder", 99)


questions_sorted = sorted(
    ds["questions"],
    key=lambda q: (stream_order(q["streamId"]), sub_order(q["subProcessId"]), q["displayOrder"]),
)

row = 2
for i, q in enumerate(questions_sorted, start=1):
    stream = stream_by_id.get(q["streamId"], {})
    sub = sub_by_id.get(q["subProcessId"], {})
    cells = [
        i,
        q["id"],
        q.get("status"),
        q.get("format", "decision"),
        q.get("flag") or "",
        stream.get("name", q["streamId"]),
        sub.get("name", q["subProcessId"]),
        ", ".join(q["scopeItemRefs"]) if q["scopeItemRefs"] else "(stream-level — empty refs)",
        len(q["scopeItemRefs"]),
        q.get("sapArea") or "",
        q.get("sapTopic") or "",
        q.get("sscuiRef") or "",
        q.get("sourceQuestionnaire") or "",
        q.get("placementBasis") or "",
        q.get("sapVerbatim") or "",
        q.get("plainLanguageSuggested") or "",
        q.get("consultantWording") or "",
        q.get("aboutText") or "",
        q.get("standardMeans") or "",
        q.get("statusNote") or "",
        "yes" if q.get("isCustom") else "",
    ]
    for c, v in enumerate(cells, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = WRAP
    row += 1

autosize(ws, [
    (1, 5),  (2, 10), (3, 12), (4, 14), (5, 18), (6, 22), (7, 28),
    (8, 28), (9, 8),  (10, 22), (11, 24), (12, 16), (13, 18), (14, 22),
    (15, 60), (16, 60), (17, 60), (18, 60), (19, 60), (20, 30), (21, 12),
])


# ─── Sheet 2: Scope items ──────────────────────────────────────────

ws = wb.create_sheet("Scope items")
columns = [
    "#",
    "Scope ID",
    "Description",
    "Value stream",
    "Sub-process",
    "SAP Business Area",
    "Has BDC coverage",
    "Question count (scope-ref join)",
    "Placement review flag",
    "Has MY process flow",
    "MY mandatory step count",
]
write_header(ws, 1, columns)
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

scope_sorted = sorted(
    ds["scopeItems"],
    key=lambda s: (stream_order(s["streamId"]), sub_order(s["subProcessId"]), s["id"]),
)
row = 2
for i, s in enumerate(scope_sorted, start=1):
    stream = stream_by_id.get(s["streamId"], {})
    sub = sub_by_id.get(s["subProcessId"], {})
    flow = flow_by_id.get(s["id"])
    cells = [
        i,
        s["id"],
        s["description"],
        stream.get("name", s["streamId"]),
        sub.get("name", s["subProcessId"]),
        s.get("sapBusinessArea") or "",
        "yes" if s["hasBdcCoverage"] else "no",
        len(questions_by_scope_ref.get(s["id"], [])),
        s.get("placementReviewFlag") or "",
        "yes" if flow else "no",
        flow["myStepCount"] if flow else 0,
    ]
    for c, v in enumerate(cells, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = WRAP
    row += 1
autosize(ws, [(1, 5), (2, 10), (3, 60), (4, 22), (5, 28), (6, 28), (7, 12), (8, 12), (9, 16), (10, 12), (11, 12)])


# ─── Sheet 3: Sub-processes ────────────────────────────────────────

ws = wb.create_sheet("Sub-processes")
columns = ["#", "Sub-process ID", "Name", "Value stream", "Type", "Display order", "Scope items", "Questions"]
write_header(ws, 1, columns)
ws.freeze_panes = "B2"

sub_sorted = sorted(
    ds["subProcesses"],
    key=lambda sp: (stream_order(sp["streamId"]), sp["displayOrder"]),
)
scope_count_by_sub = Counter(s["subProcessId"] for s in ds["scopeItems"])

row = 2
for i, sp in enumerate(sub_sorted, start=1):
    stream = stream_by_id.get(sp["streamId"], {})
    cells = [
        i,
        sp["id"],
        sp["name"],
        stream.get("name", sp["streamId"]),
        sp.get("type", ""),
        sp["displayOrder"],
        scope_count_by_sub.get(sp["id"], 0),
        len(questions_by_sub.get(sp["id"], [])),
    ]
    for c, v in enumerate(cells, start=1):
        ws.cell(row=row, column=c, value=v).alignment = WRAP
    row += 1
autosize(ws, [(1, 5), (2, 42), (3, 36), (4, 22), (5, 12), (6, 12), (7, 12), (8, 12)])


# ─── Sheet 4: Value streams ────────────────────────────────────────

ws = wb.create_sheet("Value streams")
columns = ["#", "Stream ID", "Name", "Foundation?", "Display order", "Sub-processes", "Scope items", "Questions"]
write_header(ws, 1, columns)
ws.freeze_panes = "B2"

stream_sorted = sorted(ds["valueStreams"], key=lambda s: s["displayOrder"])
sub_count = Counter(sp["streamId"] for sp in ds["subProcesses"])
scope_count = Counter(s["streamId"] for s in ds["scopeItems"])
q_count = Counter(q["streamId"] for q in ds["questions"])

row = 2
for i, s in enumerate(stream_sorted, start=1):
    cells = [
        i,
        s["id"],
        s["name"],
        "yes" if s["isFoundation"] else "no",
        s["displayOrder"],
        sub_count.get(s["id"], 0),
        scope_count.get(s["id"], 0),
        q_count.get(s["id"], 0),
    ]
    for c, v in enumerate(cells, start=1):
        ws.cell(row=row, column=c, value=v).alignment = WRAP
    row += 1
autosize(ws, [(1, 5), (2, 24), (3, 30), (4, 12), (5, 12), (6, 14), (7, 12), (8, 12)])


# ─── Sheet 5: Process flow (per-step) ──────────────────────────────

ws = wb.create_sheet("Process flow")
columns = [
    "#",
    "Scope ID",
    "Scope description",
    "Value stream",
    "Step #",
    "Activity (SAP-verbatim)",
    "Fiori app(s)",
]
write_header(ws, 1, columns)
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

row = 2
i = 0
for scope_id in sorted(steps_by_scope.keys()):
    scope = scope_by_id.get(scope_id, {})
    stream = stream_by_id.get(scope.get("streamId", ""), {})
    for step in steps_by_scope[scope_id]:
        i += 1
        cells = [
            i,
            scope_id,
            scope.get("description", ""),
            stream.get("name", scope.get("streamId", "")),
            step["stepNumber"],
            step["activity"],
            " / ".join(step["fioriApps"]) if step["fioriApps"] else "",
        ]
        for c, v in enumerate(cells, start=1):
            ws.cell(row=row, column=c, value=v).alignment = WRAP
        row += 1
autosize(ws, [(1, 6), (2, 10), (3, 50), (4, 22), (5, 8), (6, 60), (7, 50)])


# ─── Sheet 6: Per-scope flow summary ───────────────────────────────

ws = wb.create_sheet("Flow summary")
columns = [
    "#",
    "Scope ID",
    "Description",
    "Value stream",
    "Activities (raw)",
    "MY mandatory steps",
    "Optional steps",
    "Step names (→ separated)",
]
write_header(ws, 1, columns)
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

flow_sorted = sorted(
    pf["flows"],
    key=lambda f: (
        stream_order(scope_by_id.get(f["scopeItemId"], {}).get("streamId", "")),
        f["scopeItemId"],
    ),
)
row = 2
for i, f in enumerate(flow_sorted, start=1):
    scope = scope_by_id.get(f["scopeItemId"], {})
    stream = stream_by_id.get(scope.get("streamId", ""), {})
    steps = steps_by_scope.get(f["scopeItemId"], [])
    cells = [
        i,
        f["scopeItemId"],
        scope.get("description", ""),
        stream.get("name", ""),
        f["activityCount"],
        f["myStepCount"],
        f["optionalCount"],
        "  →  ".join(s["activity"] for s in steps),
    ]
    for c, v in enumerate(cells, start=1):
        ws.cell(row=row, column=c, value=v).alignment = WRAP
    row += 1
autosize(ws, [(1, 6), (2, 10), (3, 50), (4, 22), (5, 14), (6, 14), (7, 14), (8, 80)])


# ─── Sheet 7: UI defaults & placeholders ───────────────────────────
# Every default/placeholder/eyebrow string the cards render. The list
# is authored by hand against the component source — it is the
# checklist the consultant uses to verify the UI presents every data
# field correctly.

ws = wb.create_sheet("UI defaults")
columns = [
    "Surface",
    "Element",
    "Type",
    "Visible text / behaviour",
    "Source data field (if any)",
    "Notes",
]
write_header(ws, 1, columns)
ws.freeze_panes = "A2"

ui_rows = [
    # Client view — DecisionCard
    ("Client / Decision card", "Header — category", "label", "{question.sapTopic} (falls back to 'General')", "AffirmQuestion.sapTopic", "Shown left of format badge."),
    ("Client / Decision card", "Header — format badge", "static", "Decision (blue pill, bg-decision-configure/15)", "—", "Always present on Decision cards."),
    ("Client / Decision card", "Header — DEFAULT badge", "conditional", "default (teal pill) — shown only when current choice is 'standard'", "AffirmResponse.choice == 'standard'", "Visually marks the un-touched default."),
    ("Client / Decision card", "Header — consultant-added", "conditional", "consultant-added (navy pill) — shown when question.isCustom", "AffirmQuestion.isCustom", "Locked-on marker."),
    ("Client / Decision card", "Header — Q ID", "label", "{question.id} (mono right-aligned)", "AffirmQuestion.id", "Always shown."),
    ("Client / Decision card", "About line", "conditional", "**What this is about:** {aboutText}", "AffirmQuestion.aboutText", "Hidden when null."),
    ("Client / Decision card", "Main wording", "label", "{consultantWording ?? plainLanguageSuggested ?? sapVerbatim}", "AffirmQuestion.consultantWording (fallback chain)", "The client-facing prompt."),
    ("Client / Decision card", "ProcessFlowStrip (embedded)", "component", "'Where this sits · {scopeId} {description}' eyebrow + numbered steps + Fiori apps", "AffirmProcessStep[]", "One strip per scope-item ref. Renders 'no flow available' for the 17 scope items without an MY flow."),
    ("Client / Decision card", "Standard-means box", "conditional", "**What \"Adopt SAP standard\" means here** + {standardMeans}", "AffirmQuestion.standardMeans", "Hidden when null. Filled by consultant in editor."),
    ("Client / Decision card", "Verbatim toggle", "behaviour", "'Show the exact SAP wording' — chevron rotates 90° on click", "—", "Default state = collapsed."),
    ("Client / Decision card", "Verbatim box (when open)", "conditional", "'SAP verbatim · source of truth · never changed' label + {sapVerbatim} + 'SSCUI: {sscuiRef}'", "AffirmQuestion.sapVerbatim + sscuiRef", "SSCUI line hidden when sscuiRef ∈ {null, '-', 'N/A'}."),
    ("Client / Decision card", "Source line", "label", "'SAP BDC · Level 2 · {sapArea} · {sscuiRef}'", "AffirmQuestion.sapArea + sscuiRef", "Mono ink-muted."),
    ("Client / Decision card", "Choice — Standard", "button", "'Adopt SAP standard' (teal) + helper 'Keep SAP's pre-delivered setup.'", "AffirmResponse.choice='standard'", "Default selection."),
    ("Client / Decision card", "Choice — Discuss", "button", "'Discuss in workshop' (blue) + helper 'Unsure — raise it in the workshop.'", "AffirmResponse.choice='discuss'", "—"),
    ("Client / Decision card", "Choice — Deviate", "button", "'We do this differently' (amber) + helper 'You have a specific requirement.'", "AffirmResponse.choice='deviate'", "Reveals reason textarea on click."),
    ("Client / Decision card", "Reason textarea", "conditional", "Label 'Why we differ' + textarea (placeholder 'Required for a deviation.')", "AffirmResponse.reason", "Shown only when choice='deviate'. Server rejects empty reason on submit."),
    # Client view — InfoCard
    ("Client / Information card", "Format badge", "static", "'Information' (ink-tint pill)", "—", "Always present on Information cards."),
    ("Client / Information card", "About line", "conditional", "**What this is about:** {aboutText}", "AffirmQuestion.aboutText", "Same rule as Decision."),
    ("Client / Information card", "Main wording", "label", "{consultantWording ?? plainLanguageSuggested ?? sapVerbatim}", "AffirmQuestion.consultantWording", "—"),
    ("Client / Information card", "Answer textarea", "input", "Label 'Your answer' + textarea (placeholder 'Type your answer here — e.g. list the items, describe your setup…')", "AffirmResponse.reason", "Free-text answer; 2000 char cap."),
    ("Client / Information card", "ProcessFlowStrip (embedded)", "component", "Same as Decision card — between wording and answer textarea", "AffirmProcessStep[]", "—"),
    ("Client / Information card", "Flag toggle", "button", "'Flag for workshop discussion' / 'Flagged for workshop' (blue when on) + helper 'Use if you would rather cover this live.'", "AffirmResponse.choice='discuss'", "No 'Adopt SAP standard' on Information cards — by design."),
    ("Client / Information card", "Verbatim toggle", "behaviour", "Same as Decision card", "AffirmQuestion.sapVerbatim", "—"),
    # Client view — top section
    ("Client / Page", "Eyebrow", "static", "'Fit-to-Standard pre-workshop affirmation'", "—", "Above the page title."),
    ("Client / Page", "Title", "label", "'{bundle.client} · {streamName}'", "AffirmBundle.client", "Stream name = single stream name, or 'N value streams'."),
    ("Client / Page", "Stat strip", "live", "Questions · Adopt standard · Discuss · We differ · Information (counts)", "AffirmResponse[] aggregated", "Decision split + Information count."),
    ("Client / Page", "Progress bar", "live", "Stacked bar: std / cfg / cust % of Decision responses", "AffirmResponse[]", "Decision answers only — Information excluded from %."),
    ("Client / Page", "How to answer", "static", "Card explaining Decision vs Information. Always shown.", "—", "v2 §7."),
    ("Client / Page", "Submit button", "static", "'Submit my affirmation' (CTA red)", "—", "Sticky bottom release bar."),
    # Consultant — Editor row
    ("Consultant / Editor row", "Header — category", "label", "{sapTopic ?? sapArea ?? 'General'}", "AffirmQuestion.sapTopic|sapArea", "Same fallback chain as client header."),
    ("Consultant / Editor row", "Header — format badge", "live", "'Decision' (blue) / 'Information' (gray) per current join-row format", "AffirmBundleQuestion.format", "—"),
    ("Consultant / Editor row", "Header — consultant-added", "conditional", "'consultant-added' (navy)", "AffirmQuestion.isCustom", "—"),
    ("Consultant / Editor row", "Header — flagged", "conditional", "'flagged' (amber) when question.flag='config-how-to'", "AffirmQuestion.flag", "13 questions flagged."),
    ("Consultant / Editor row", "Header — SAP-excluded", "conditional", "'SAP-excluded' (ink-tint) when status='excluded'", "AffirmQuestion.status", "15 excluded rows ship pre-disabled."),
    ("Consultant / Editor row", "Header — Q ID", "label", "{question.id}", "—", "Mono right-aligned."),
    ("Consultant / Editor row", "Plain-language wording", "editable", "Label 'Plain-language wording (editable)' + textarea — saves on blur", "AffirmQuestion.consultantWording", "Replaces what the client sees."),
    ("Consultant / Editor row", "About-text input", "editable", "Label 'About this question' + textarea (placeholder 'What this question is about — shown above the wording on the client card.')", "AffirmQuestion.aboutText", "Seeded from Standard-Answer-Layer Topic Definition."),
    ("Consultant / Editor row", "Standard-means input", "editable (Decision only)", "Label 'What \"Adopt SAP standard\" means here' + textarea (placeholder 'Plain-language translation of the SAP standard — shown in the teal box on the client card.')", "AffirmQuestion.standardMeans", "Hidden on Information format."),
    ("Consultant / Editor row", "ProcessFlowStrip preview", "component", "Same embedded strip the client will see — one per scope-item ref", "AffirmProcessStep[]", "Read-only preview."),
    ("Consultant / Editor row", "SAP verbatim (locked)", "disclosure", "<details> summary 'SAP verbatim (locked) — never editable' + box", "AffirmQuestion.sapVerbatim", "Hidden by default. Verbatim never editable."),
    ("Consultant / Editor row", "Format toggle", "control", "Decision / Information pill toggle (navy = on)", "AffirmBundleQuestion.format", "Persists to join row, not the question bank."),
    ("Consultant / Editor row", "Reorder ↑ / ↓", "control", "Move row up/down within sub-process; swaps displayOrder with neighbour", "AffirmBundleQuestion.displayOrder", "Disabled at list ends."),
    ("Consultant / Editor row", "Enable / Disable", "control", "Toggle button — 'Disable' (default) / 'Enable' (cta colour when disabled)", "AffirmBundleQuestion.enabled", "Disabled rows stay in the record but never reach the client view."),
    ("Consultant / Editor row", "Delete (custom only)", "control", "Delete button — confirm modal — only on isCustom=true", "—", "SAP questions return 422 if DELETE attempted."),
    # Consultant — Editor page
    ("Consultant / Editor page", "Stat strip", "live", "In bundle / Decision / Information / Excluded / Custom (counts)", "AffirmBundleQuestion[] aggregated", "—"),
    ("Consultant / Editor page", "Add custom question", "control", "Per-sub-process '+ Add custom question' opens inline form (wording, aboutText, standardMeans, format)", "POST /api/affirm/bundles/[id]/questions", "Adds row to that sub-process. Marked consultant-added forever."),
    ("Consultant / Editor page", "Issue CTA (draft only)", "control", "Sticky bottom 'Issue to client' button. Disabled when 0 enabled questions.", "POST /api/affirm/bundles/[id]/issue", "Transitions DRAFT → ISSUED."),
    # Tree picker
    ("Consultant / Scope picker", "Tier labels", "static", "Tier 1 — value stream / Tier 2 — sub-process / Tier 3 — scope item", "—", "Above the tree."),
    ("Consultant / Scope picker", "Stream row", "control", "Chevron (expand only). No checkbox at stream level.", "—", "v2 §1."),
    ("Consultant / Scope picker", "Sub-process row", "control", "Tri-state checkbox (Select all / Clear) + coverage chip ('N questions' or 'no affirm-set')", "AffirmScopeItem.hasBdcCoverage", "v2 §3, §5."),
    ("Consultant / Scope picker", "Scope item row", "control", "Checkbox + scope code (mono) + description + 'N questions' or 'no affirm-set' chip + optional 'pending review'", "AffirmScopeItem.placementReviewFlag", "63 items carry pending-curation."),
    ("Consultant / Scope picker", "Side panel CTAs", "control", "'Create bundle' / 'Save scope' → routes to /affirm/[id]/questions", "—", "Editor is next step."),
    # Review screen
    ("Consultant / Review screen", "Lifecycle stepper", "live", "Draft → Issued → Submitted → Released. Current state highlighted; previous ticked.", "AffirmBundle.state", "v2 §6."),
    ("Consultant / Review screen", "Badge legend", "static", "Defines every chip used across the surface", "—", "v2 §6."),
    ("Consultant / Review screen", "Release CTA (submitted only)", "control", "Sticky bottom 'Release' (CTA red) — opens confirm modal", "POST /api/affirm/bundles/[id]/release", "Two-step gate."),
    # Output
    ("Consultant / Output", "Doc rows", "static", "Pre-workshop pack PDF · Workshop agenda DOCX · Deviation register XLSX · Audit trail CSV", "—", "Generation pending — UI lists artifacts but manual send."),
]

for r in ui_rows:
    row_idx = ws.max_row + 1
    for c, v in enumerate(r, start=1):
        ws.cell(row=row_idx, column=c, value=v).alignment = WRAP
autosize(ws, [(1, 30), (2, 28), (3, 14), (4, 80), (5, 32), (6, 60)])


# ─── Save ───────────────────────────────────────────────────────────

wb.save(OUT)
print(f"wrote {OUT}")
print("sheets:")
for s in wb.sheetnames:
    sheet = wb[s]
    print(f"  - {s:<18}  rows={sheet.max_row:>5}  cols={sheet.max_column}")
